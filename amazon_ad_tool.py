"""
亚马逊广告诊断工具 v1.0
============================
双击运行，选择你的亚马逊广告报告（搜索词报告 / 推广商品报告），
自动生成诊断分析Excel，无需联网。

依赖：pip install pandas openpyxl
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, sys, threading, datetime

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    HAS_TK = True
except ImportError:
    HAS_TK = False

# ══════════════════════════════════════════════════════════════
# 诊断阈值（你可以根据自己产品修改这里的数字）
# ══════════════════════════════════════════════════════════════
ACOS_HIGH = 0.40       # ACoS 超过此值 → 高ACoS
ACOS_LOW  = 0.20       # ACoS 低于此值 → 优质词
CTR_LOW   = 0.003      # CTR 低于 0.3% → 低CTR
SPEND_THRESHOLD = 5    # 花费超过 $5 且 0单 → 烧钱词
CLICK_NO_SALE   = 8    # 点击超过 8 次且 0单 → 点击无转化

# ══════════════════════════════════════════════════════════════
# 样式常量
# ══════════════════════════════════════════════════════════════
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
NORMAL_FONT = Font(name='Arial', size=10)
TITLE_FONT  = Font(name='Arial', bold=True, color='1F4E79', size=14)
METRIC_FONT = Font(name='Arial', bold=True, color='1F4E79', size=18)
LABEL_FONT  = Font(name='Arial', color='666666', size=9)
RED_FILL    = PatternFill('solid', fgColor='FDE8E8')
GREEN_FILL  = PatternFill('solid', fgColor='E2EFDA')
YELLOW_FILL = PatternFill('solid', fgColor='FFF2CC')
ORANGE_FILL = PatternFill('solid', fgColor='FCE4D6')
BLUE_FILL   = PatternFill('solid', fgColor='DAEEF3')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))


# ══════════════════════════════════════════════════════════════
# 工具函数
# ══════════════════════════════════════════════════════════════
def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_rows(ws, r1, r2, ncol):
    for r in range(r1, r2 + 1):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c)
            cell.font, cell.border = NORMAL_FONT, THIN_BORDER
            cell.alignment = Alignment(vertical='center')

def auto_w(ws, ncol):
    for c in range(1, ncol + 1):
        ml = max((len(str(ws.cell(row=r, column=c).value or ''))
                  for r in range(1, min(ws.max_row + 1, 60))), default=8)
        ws.column_dimensions[get_column_letter(c)].width = min(max(ml + 3, 10), 35)

def fmt_pct(ws, col, r1, r2):
    for r in range(r1, r2 + 1):
        ws.cell(row=r, column=col).number_format = '0.00%'

def fmt_money(ws, col, r1, r2):
    for r in range(r1, r2 + 1):
        ws.cell(row=r, column=col).number_format = '$#,##0.00'

def fmt_num(ws, col, r1, r2):
    for r in range(r1, r2 + 1):
        ws.cell(row=r, column=col).number_format = '#,##0'


# ══════════════════════════════════════════════════════════════
# 数据加载 & 列名映射（支持中英文亚马逊后台报告）
# ══════════════════════════════════════════════════════════════
CN_MAP = {
    '广告活动名称': 'Campaign', '广告组名称': 'AdGroup',
    '投放': 'Targeting', '匹配类型': 'MatchType',
    '客户搜索词': 'SearchTerm', '展示量': 'Impressions',
    '点击量': 'Clicks', '点击率 (CTR)': 'CTR',
    '单次点击成本 (CPC)': 'CPC', '花费': 'Spend',
    '7天总销售额': 'Sales', '广告投入产出比 (ACOS) 总计': 'ACoS',
    '总广告投资回报率 (ROAS)': 'ROAS', '7天总订单数(#)': 'Orders',
    '7天总销售量(#)': 'Units', '7天的转化率': 'CVR',
    '7天内广告SKU销售量(#)': 'SKU_Units', '7天内其他SKU销售量(#)': 'Other_Units',
    '7天内广告SKU销售额': 'SKU_Sales', '7天内其他SKU销售额': 'Other_Sales',
    '广告SKU': 'SKU', '广告ASIN': 'ASIN', '广告组合名称': 'Portfolio',
    # English column names
    'Campaign Name': 'Campaign', 'Ad Group Name': 'AdGroup',
    'Targeting': 'Targeting', 'Match Type': 'MatchType',
    'Customer Search Term': 'SearchTerm', 'Impressions': 'Impressions',
    'Clicks': 'Clicks', 'Click-Thru Rate (CTR)': 'CTR',
    'Cost Per Click (CPC)': 'CPC', 'Spend': 'Spend',
    '7 Day Total Sales': 'Sales', 'Total Advertising Cost of Sales (ACoS)': 'ACoS',
    'Total Return on Advertising Spend (ROAS)': 'ROAS',
    '7 Day Total Orders (#)': 'Orders', '7 Day Total Units (#)': 'Units',
    '7 Day Conversion Rate': 'CVR',
    '14 Day Total Sales': 'Sales', '14 Day Total Orders (#)': 'Orders',
    '14 Day Total Units (#)': 'Units',
    'Advertised SKU': 'SKU', 'Advertised ASIN': 'ASIN',
    'Portfolio name': 'Portfolio',
}


def load_file(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == '.csv':
        # try utf-8 first, then gbk, then latin-1
        for enc in ['utf-8', 'gbk', 'utf-8-sig', 'latin-1']:
            try:
                df = pd.read_csv(path, encoding=enc)
                break
            except (UnicodeDecodeError, Exception):
                continue
        else:
            raise ValueError(f"无法读取文件编码: {path}")
    else:
        df = pd.read_excel(path)

    rename = {c: CN_MAP[c] for c in df.columns if c in CN_MAP}
    df = df.rename(columns=rename)

    nums = ['Impressions', 'Clicks', 'CTR', 'CPC', 'Spend', 'Orders', 'Units',
            'Sales', 'ACoS', 'ROAS', 'CVR', 'SKU_Units', 'Other_Units',
            'SKU_Sales', 'Other_Sales']
    for c in nums:
        if c in df.columns:
            df[c] = pd.to_numeric(
                df[c].astype(str).str.replace('%', '').str.replace('$', '').str.replace(',', ''),
                errors='coerce')

    for c in ['CTR', 'ACoS', 'CVR']:
        if c in df.columns and df[c].dropna().median() > 1:
            df[c] /= 100

    if 'ACoS' not in df.columns or df['ACoS'].isna().all():
        df['ACoS'] = df['Spend'] / df['Sales'].replace(0, np.nan)
    if 'CVR' not in df.columns or df['CVR'].isna().all():
        df['CVR'] = df['Orders'] / df['Clicks'].replace(0, np.nan)
    if 'Units' not in df.columns:
        df['Units'] = df.get('Orders', 0)

    return df


def detect_report_type(df):
    """判断是搜索词报告还是推广商品报告"""
    cols = set(df.columns)
    if 'SearchTerm' in cols:
        return 'search_term'
    elif 'ASIN' in cols or 'SKU' in cols:
        return 'product'
    else:
        return 'search_term'  # 默认当搜索词报告


# ══════════════════════════════════════════════════════════════
# 关键词诊断分类
# ══════════════════════════════════════════════════════════════
def classify(df):
    labels, notes, pris = [], [], []
    for _, r in df.iterrows():
        sp = r.get('Spend', 0) or 0
        cl = r.get('Clicks', 0) or 0
        od = r.get('Orders', 0) or 0
        imp = r.get('Impressions', 0) or 0
        acos = r.get('ACoS', None)
        ctr = r.get('CTR', None)
        if pd.isna(acos): acos = None
        if pd.isna(ctr): ctr = None
        lb, nt, pri = [], [], '—'

        if sp >= SPEND_THRESHOLD and od == 0:
            lb.append('🔴 烧钱词')
            nt.append(f'花费${sp:.2f}，0订单')
            pri = '‼️ 紧急'
        elif acos and acos > ACOS_HIGH and od > 0:
            lb.append('🟠 高ACoS')
            nt.append(f'ACoS={acos:.0%}')
            pri = '⚠️ 高'

        if cl >= CLICK_NO_SALE and od == 0:
            lb.append('🔴 点击无转化')
            nt.append(f'{cl}次点击0单')
            if pri == '—':
                pri = '‼️ 紧急'

        if acos and 0 < acos <= ACOS_LOW and od >= 2:
            lb.append('🟢 优质词')
            nt.append(f'ACoS={acos:.0%},{od}单')
            pri = '🌟 加投'
        elif acos and ACOS_LOW < acos <= ACOS_HIGH and od >= 1:
            lb.append('🟡 潜力词')
            nt.append(f'ACoS={acos:.0%}可优化')
            if pri == '—':
                pri = '📌 关注'

        if ctr is not None and ctr < CTR_LOW and imp > 500:
            lb.append('🔵 低CTR')
            nt.append(f'CTR={ctr:.2%}')

        if sp >= SPEND_THRESHOLD and od == 0 and cl >= 3:
            if '🔴 烧钱词' not in lb:
                lb.append('❌ 建议否定')
            if '建议加否定' not in ''.join(nt):
                nt.append('建议加否定')

        labels.append(' | '.join(lb) if lb else '—')
        notes.append('；'.join(nt) if nt else '正常')
        pris.append(pri)

    df['诊断标签'] = labels
    df['诊断说明'] = notes
    df['优先级'] = pris
    return df


# ══════════════════════════════════════════════════════════════
# 写入 Sheet
# ══════════════════════════════════════════════════════════════
def write_sheet(ws, df, cols, hdr_row, color_col='诊断标签'):
    for c, h in enumerate(cols, 1):
        ws.cell(row=hdr_row, column=c, value=h)
    style_header(ws, hdr_row, len(cols))

    for ri, (_, rd) in enumerate(df.iterrows(), hdr_row + 1):
        for ci, cn in enumerate(cols, 1):
            v = rd.get(cn, '')
            if pd.isna(v):
                v = '—'
            ws.cell(row=ri, column=ci, value=v)
        if color_col and color_col in rd.index:
            tag = str(rd.get(color_col, ''))
            fill = None
            if '烧钱' in tag or '点击无转化' in tag or '亏损' in tag:
                fill = RED_FILL
            elif '优质' in tag or '盈利' in tag:
                fill = GREEN_FILL
            elif '潜力' in tag or '可优化' in tag:
                fill = YELLOW_FILL
            elif '高ACoS' in tag:
                fill = ORANGE_FILL
            elif '低CTR' in tag:
                fill = BLUE_FILL
            if fill:
                for ci in range(1, len(cols) + 1):
                    ws.cell(row=ri, column=ci).fill = fill

    er = hdr_row + len(df)
    style_rows(ws, hdr_row + 1, er, len(cols))
    for ci, cn in enumerate(cols, 1):
        if cn in ['Spend', 'Sales', 'CPC', 'SKU_Sales', 'Other_Sales']:
            fmt_money(ws, ci, hdr_row + 1, er)
        elif cn in ['ACoS', 'CTR', 'CVR']:
            fmt_pct(ws, ci, hdr_row + 1, er)
        elif cn in ['Impressions', 'Clicks', 'Orders', 'Units', '问题词数']:
            fmt_num(ws, ci, hdr_row + 1, er)
    auto_w(ws, len(cols))
    return er


# ══════════════════════════════════════════════════════════════
# 生成诊断报告（核心函数）
# ══════════════════════════════════════════════════════════════
def build_report(df_st, df_prod, output_path, progress_callback=None):
    def log(msg):
        if progress_callback:
            progress_callback(msg)

    wb = Workbook()

    # 预处理搜索词
    df_all = df_st.copy()
    df_kw = df_st[df_st.get('SearchTerm', pd.Series([''])) != '*'].copy()
    df_star = df_st[df_st.get('SearchTerm', pd.Series([''])) == '*'].copy()

    log("正在诊断关键词...")
    df_kw = classify(df_kw)
    if len(df_star) > 0:
        df_star['诊断标签'] = '—'
        df_star['诊断说明'] = '自动投放汇总'
        df_star['优先级'] = '—'
    df_full = pd.concat([df_kw, df_star], ignore_index=True)

    # 统计
    total_kw = len(df_kw)
    burn = len(df_kw[df_kw['诊断标签'].str.contains('烧钱词', na=False)])
    high_a = len(df_kw[df_kw['诊断标签'].str.contains('高ACoS', na=False)])
    good = len(df_kw[df_kw['诊断标签'].str.contains('优质词', na=False)])
    pot = len(df_kw[df_kw['诊断标签'].str.contains('潜力词', na=False)])
    neg = burn  # 烧钱词 = 建议否定
    low_c = len(df_kw[df_kw['诊断标签'].str.contains('低CTR', na=False)])
    waste = df_kw[df_kw['诊断标签'].str.contains('烧钱词', na=False)]['Spend'].sum()

    # ─── Sheet 1: 总览仪表盘 ───
    log("生成总览仪表盘...")
    ws = wb.active
    ws.title = '总览仪表盘'
    ws.sheet_properties.tabColor = '1F4E79'
    ws.merge_cells('A1:I1')
    ws['A1'] = '亚马逊广告全面诊断报告'
    ws['A1'].font = TITLE_FONT
    ws.row_dimensions[1].height = 40

    ts = df_all['Spend'].sum()
    tsa = df_all['Sales'].sum()
    to_ = df_all['Orders'].sum()
    tc = df_all['Clicks'].sum()
    ti = df_all['Impressions'].sum()
    acos_o = ts / tsa if tsa > 0 else 0
    roas_o = tsa / ts if ts > 0 else 0
    ctr_o = tc / ti if ti > 0 else 0
    cvr_o = to_ / tc if tc > 0 else 0
    cpc_o = ts / tc if tc > 0 else 0

    kpis = [
        ('总花费', f'${ts:,.2f}'), ('总销售额', f'${tsa:,.2f}'),
        ('总订单数', f'{to_:,}'), ('总点击', f'{tc:,}'),
        ('整体ACoS', f'{acos_o:.1%}'), ('整体ROAS', f'{roas_o:.2f}'),
        ('平均CTR', f'{ctr_o:.2%}'), ('平均CVR', f'{cvr_o:.1%}'),
        ('平均CPC', f'${cpc_o:.2f}')
    ]
    for i, (label, value) in enumerate(kpis):
        row = 3 if i < 5 else 6
        col = (i % 5) * 2 + 1
        ws.cell(row=row, column=col, value=label).font = LABEL_FONT
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=row + 1, column=col, value=value).font = METRIC_FONT
        ws.cell(row=row + 1, column=col).alignment = Alignment(horizontal='center')

    # 健康分布
    r0 = 9
    ws.cell(row=r0, column=1, value='搜索词健康分布').font = Font(name='Arial', bold=True, size=12, color='1F4E79')
    for c, h in enumerate(['类型', '数量', '占比', '涉及花费'], 1):
        ws.cell(row=r0 + 1, column=c, value=h)
    style_header(ws, r0 + 1, 4)

    dist = [
        ('🔴 烧钱词（高花费0转化）', burn, RED_FILL, '烧钱词'),
        ('🟠 高ACoS（>40%）', high_a, ORANGE_FILL, '高ACoS'),
        ('🟢 优质词（ACoS<20%）', good, GREEN_FILL, '优质词'),
        ('🟡 潜力词（20%-40%）', pot, YELLOW_FILL, '潜力词'),
        ('🔵 低CTR词', low_c, BLUE_FILL, '低CTR'),
    ]
    for i, (name, cnt, fill, tag_key) in enumerate(dist):
        r = r0 + 2 + i
        ws.cell(row=r, column=1, value=name).fill = fill
        ws.cell(row=r, column=2, value=cnt)
        ws.cell(row=r, column=3, value=cnt / total_kw if total_kw > 0 else 0)
        ws.cell(row=r, column=3).number_format = '0.0%'
        w = df_kw[df_kw['诊断标签'].str.contains(tag_key, na=False)]['Spend'].sum()
        ws.cell(row=r, column=4, value=w)
        ws.cell(row=r, column=4).number_format = '$#,##0.00'
        for c in range(1, 5):
            ws.cell(row=r, column=c).font = NORMAL_FONT
            ws.cell(row=r, column=c).border = THIN_BORDER

    # 优化建议
    sr = r0 + 2 + len(dist) + 1
    ws.cell(row=sr, column=1, value='💡 核心优化建议').font = Font(name='Arial', bold=True, size=12, color='1F4E79')
    sugs = []
    if burn > 0:
        sugs.append(f'① 立即否定 {burn} 个烧钱词，可节省约 ${waste:,.2f}')
    if high_a > 0:
        ha_sp = df_kw[df_kw['诊断标签'].str.contains('高ACoS', na=False)]['Spend'].sum()
        sugs.append(f'② {high_a} 个高ACoS词（花费${ha_sp:,.2f}），降低竞价或暂停')
    if good > 0:
        gs = df_kw[df_kw['诊断标签'].str.contains('优质词', na=False)]['Sales'].sum()
        sugs.append(f'③ {good} 个优质词贡献${gs:,.2f}销售，建议提高竞价抢量')
    if pot > 0:
        sugs.append(f'④ {pot} 个潜力词可通过优化listing和竞价提升转化')
    if low_c > 0:
        sugs.append(f'⑤ {low_c} 个低CTR词需检查主图/标题/价格')
    for i, s in enumerate(sugs):
        ws.cell(row=sr + 1 + i, column=1, value=s).font = Font(name='Arial', size=10, color='333333')
        ws.merge_cells(start_row=sr + 1 + i, start_column=1, end_row=sr + 1 + i, end_column=8)
    auto_w(ws, 10)

    # ─── Sheet 2: 烧钱词 ───
    log("生成烧钱词清单...")
    ws2 = wb.create_sheet('烧钱词&否定词')
    ws2.sheet_properties.tabColor = 'CC0000'
    prob = df_kw[df_kw['诊断标签'].str.contains('烧钱词|点击无转化|高ACoS', na=False)].copy()
    prob = prob.sort_values('Spend', ascending=False)
    cols2 = [c for c in ['SearchTerm','Campaign','AdGroup','MatchType','Impressions','Clicks',
             'Spend','Orders','Sales','ACoS','CTR','CVR','诊断标签','优先级'] if c in prob.columns]
    ws2.cell(row=1, column=1, value=f'问题搜索词（共{len(prob)}个）— 按花费排列').font = \
        Font(name='Arial', bold=True, size=12, color='CC0000')
    ws2.merge_cells(f'A1:{get_column_letter(len(cols2))}1')
    write_sheet(ws2, prob, cols2, 2)

    # ─── Sheet 3: 优质词 ───
    log("生成优质词清单...")
    ws3 = wb.create_sheet('优质词&潜力词')
    ws3.sheet_properties.tabColor = '006100'
    gd = df_kw[df_kw['诊断标签'].str.contains('优质词|潜力词', na=False)].copy()
    gd = gd.sort_values('Sales', ascending=False)
    cols3 = [c for c in ['SearchTerm','Campaign','AdGroup','MatchType','Impressions','Clicks',
             'Spend','Orders','Sales','ACoS','ROAS','CTR','CVR','诊断标签'] if c in gd.columns]
    ws3.cell(row=1, column=1, value=f'优质&潜力词（共{len(gd)}个）— 按销售额排列').font = \
        Font(name='Arial', bold=True, size=12, color='006100')
    ws3.merge_cells(f'A1:{get_column_letter(len(cols3))}1')
    write_sheet(ws3, gd, cols3, 2)

    # ─── Sheet 4: Campaign分析 ───
    log("生成Campaign分析...")
    ws4 = wb.create_sheet('Campaign分析')
    ws4.sheet_properties.tabColor = '4472C4'
    ca = df_all.groupby('Campaign').agg(
        Impressions=('Impressions', 'sum'), Clicks=('Clicks', 'sum'),
        Spend=('Spend', 'sum'), Orders=('Orders', 'sum'), Sales=('Sales', 'sum')).reset_index()
    ca['CTR'] = ca['Clicks'] / ca['Impressions'].replace(0, np.nan)
    ca['CVR'] = ca['Orders'] / ca['Clicks'].replace(0, np.nan)
    ca['ACoS'] = ca['Spend'] / ca['Sales'].replace(0, np.nan)
    ca['ROAS'] = ca['Sales'] / ca['Spend'].replace(0, np.nan)
    ca['CPC'] = ca['Spend'] / ca['Clicks'].replace(0, np.nan)
    pc = df_kw[df_kw['诊断标签'].str.contains('烧钱词', na=False)].groupby('Campaign').size().reset_index(name='问题词数')
    ca = ca.merge(pc, on='Campaign', how='left')
    ca['问题词数'] = ca['问题词数'].fillna(0).astype(int)
    ca = ca.sort_values('Spend', ascending=False)
    cols4 = ['Campaign', 'Impressions', 'Clicks', 'CTR', 'CPC', 'Spend', 'Orders', 'Sales', 'ACoS', 'ROAS', 'CVR', '问题词数']
    ws4.cell(row=1, column=1, value='Campaign维度汇总').font = Font(name='Arial', bold=True, size=12, color='1F4E79')
    ws4.merge_cells(f'A1:{get_column_letter(len(cols4))}1')
    write_sheet(ws4, ca, cols4, 2, color_col=None)

    # ─── Sheet 5: 匹配类型 ───
    log("生成匹配类型分析...")
    group_col = 'Targeting' if 'Targeting' in df_kw.columns else 'MatchType'
    if group_col in df_kw.columns:
        ws5 = wb.create_sheet('匹配类型分析')
        ws5.sheet_properties.tabColor = '7030A0'
        mt = df_kw.groupby(group_col).agg(
            Impressions=('Impressions', 'sum'), Clicks=('Clicks', 'sum'),
            Spend=('Spend', 'sum'), Orders=('Orders', 'sum'), Sales=('Sales', 'sum')).reset_index()
        mt['CTR'] = mt['Clicks'] / mt['Impressions'].replace(0, np.nan)
        mt['CVR'] = mt['Orders'] / mt['Clicks'].replace(0, np.nan)
        mt['ACoS'] = mt['Spend'] / mt['Sales'].replace(0, np.nan)
        mt['ROAS'] = mt['Sales'] / mt['Spend'].replace(0, np.nan)
        mt = mt.sort_values('Spend', ascending=False)
        cols5 = [group_col, 'Impressions', 'Clicks', 'CTR', 'Spend', 'Orders', 'Sales', 'ACoS', 'ROAS', 'CVR']
        ws5.cell(row=1, column=1, value='投放/匹配类型对比').font = Font(name='Arial', bold=True, size=12, color='7030A0')
        ws5.merge_cells(f'A1:{get_column_letter(len(cols5))}1')
        write_sheet(ws5, mt, cols5, 2, color_col=None)

    # ─── Sheet 6: 推广商品 ───
    if df_prod is not None and len(df_prod) > 0:
        log("生成推广商品分析...")
        ws6 = wb.create_sheet('推广商品分析')
        ws6.sheet_properties.tabColor = 'ED7D31'
        dp = df_prod.sort_values('Spend', ascending=False)
        dp_labels = []
        for _, r in dp.iterrows():
            a = r.get('ACoS', None)
            if pd.isna(a): a = None
            od = r.get('Orders', 0) or 0
            sp = r.get('Spend', 0) or 0
            if sp > 10 and od == 0:
                dp_labels.append('🔴 亏损')
            elif a and a > ACOS_HIGH:
                dp_labels.append('🟠 高ACoS')
            elif a and a <= ACOS_LOW and od >= 3:
                dp_labels.append('🟢 盈利')
            elif a and ACOS_LOW < a <= ACOS_HIGH:
                dp_labels.append('🟡 可优化')
            else:
                dp_labels.append('—')
        dp['商品诊断'] = dp_labels
        cols6 = [c for c in ['ASIN', 'SKU', 'Campaign', 'Impressions', 'Clicks', 'Spend',
                 'Orders', 'Sales', 'ACoS', 'ROAS', 'CTR', 'CVR', '商品诊断'] if c in dp.columns]
        ws6.cell(row=1, column=1, value=f'推广商品分析（{len(dp)}条）').font = \
            Font(name='Arial', bold=True, size=12, color='ED7D31')
        ws6.merge_cells(f'A1:{get_column_letter(len(cols6))}1')
        write_sheet(ws6, dp, cols6, 2, color_col='商品诊断')

    # ─── Sheet 7: 完整明细 ───
    log("生成完整明细...")
    ws7 = wb.create_sheet('搜索词明细')
    ws7.sheet_properties.tabColor = '808080'
    dfs = df_kw.sort_values('Spend', ascending=False)
    cols7 = [c for c in ['SearchTerm', 'Campaign', 'AdGroup', 'Targeting', 'MatchType',
             'Impressions', 'Clicks', 'CTR', 'CPC', 'Spend', 'Orders', 'Units', 'Sales',
             'ACoS', 'ROAS', 'CVR', '诊断标签', '诊断说明', '优先级'] if c in dfs.columns]
    write_sheet(ws7, dfs, cols7, 1, color_col='诊断标签')
    ws7.freeze_panes = 'A2'

    log("保存报告...")
    wb.save(output_path)

    return {
        'total': total_kw, 'burn': burn, 'high_acos': high_a,
        'good': good, 'potential': pot, 'low_ctr': low_c,
        'waste': waste, 'spend': ts, 'sales': tsa, 'orders': to_,
        'acos': acos_o
    }


# ══════════════════════════════════════════════════════════════
# GUI 界面
# ══════════════════════════════════════════════════════════════
class App:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("亚马逊广告诊断工具 v1.0")
        self.root.geometry("680x580")
        self.root.resizable(False, False)

        # 配色
        self.root.configure(bg='#F5F7FA')
        self.file_st = tk.StringVar()
        self.file_prod = tk.StringVar()

        self._build_ui()

    def _build_ui(self):
        bg = '#F5F7FA'

        # 标题
        tk.Label(self.root, text="📊 亚马逊广告诊断工具", font=('微软雅黑', 18, 'bold'),
                 fg='#1F4E79', bg=bg).pack(pady=(20, 5))
        tk.Label(self.root, text="上传报告 → 一键生成诊断Excel（完全离线，无需联网）",
                 font=('微软雅黑', 10), fg='#666', bg=bg).pack(pady=(0, 15))

        # 文件选择区
        frame = tk.Frame(self.root, bg=bg)
        frame.pack(padx=30, fill='x')

        # 搜索词报告（必选）
        tk.Label(frame, text="① 搜索词报告（必选）:", font=('微软雅黑', 10, 'bold'),
                 fg='#333', bg=bg, anchor='w').pack(fill='x', pady=(5, 2))
        f1 = tk.Frame(frame, bg=bg)
        f1.pack(fill='x', pady=(0, 10))
        tk.Entry(f1, textvariable=self.file_st, font=('微软雅黑', 9), width=50,
                 state='readonly').pack(side='left', expand=True, fill='x')
        tk.Button(f1, text="选择文件", command=self._pick_st,
                  font=('微软雅黑', 9), bg='#4472C4', fg='white',
                  relief='flat', padx=12).pack(side='right', padx=(8, 0))

        # 推广商品报告（可选）
        tk.Label(frame, text="② 推广商品报告（可选）:", font=('微软雅黑', 10, 'bold'),
                 fg='#333', bg=bg, anchor='w').pack(fill='x', pady=(5, 2))
        f2 = tk.Frame(frame, bg=bg)
        f2.pack(fill='x', pady=(0, 10))
        tk.Entry(f2, textvariable=self.file_prod, font=('微软雅黑', 9), width=50,
                 state='readonly').pack(side='left', expand=True, fill='x')
        tk.Button(f2, text="选择文件", command=self._pick_prod,
                  font=('微软雅黑', 9), bg='#4472C4', fg='white',
                  relief='flat', padx=12).pack(side='right', padx=(8, 0))

        # 阈值设置
        tk.Label(frame, text="③ 诊断阈值（可保持默认）:", font=('微软雅黑', 10, 'bold'),
                 fg='#333', bg=bg, anchor='w').pack(fill='x', pady=(10, 5))
        tf = tk.Frame(frame, bg=bg)
        tf.pack(fill='x')

        self.v_acos_high = tk.StringVar(value='40')
        self.v_acos_low = tk.StringVar(value='20')
        self.v_spend = tk.StringVar(value='5')

        for i, (label, var) in enumerate([
            ('高ACoS阈值 (%)', self.v_acos_high),
            ('优质词ACoS上限 (%)', self.v_acos_low),
            ('烧钱花费阈值 ($)', self.v_spend)
        ]):
            tk.Label(tf, text=label, font=('微软雅黑', 9), fg='#555', bg=bg).grid(row=0, column=i*2, padx=(0, 5))
            tk.Entry(tf, textvariable=var, font=('微软雅黑', 9), width=6, justify='center').grid(row=0, column=i*2+1, padx=(0, 20))

        # 开始按钮
        self.btn = tk.Button(self.root, text="🚀 开始诊断", command=self._run,
                             font=('微软雅黑', 13, 'bold'), bg='#1F4E79', fg='white',
                             relief='flat', padx=40, pady=8, cursor='hand2')
        self.btn.pack(pady=20)

        # 进度
        self.progress = ttk.Progressbar(self.root, mode='indeterminate', length=400)
        self.progress.pack(pady=(0, 5))

        self.status = tk.Label(self.root, text="等待选择文件...", font=('微软雅黑', 9),
                               fg='#888', bg=bg)
        self.status.pack()

        # 结果
        self.result_text = tk.Text(self.root, height=6, font=('Consolas', 9),
                                   bg='#FFFFFF', fg='#333', relief='flat',
                                   state='disabled', wrap='word')
        self.result_text.pack(padx=30, pady=(10, 15), fill='x')

    def _pick_st(self):
        f = filedialog.askopenfilename(
            title="选择搜索词报告",
            filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv"), ("所有文件", "*.*")])
        if f:
            self.file_st.set(f)
            self.status.config(text=f"已选择: {os.path.basename(f)}")

    def _pick_prod(self):
        f = filedialog.askopenfilename(
            title="选择推广商品报告",
            filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv"), ("所有文件", "*.*")])
        if f:
            self.file_prod.set(f)

    def _run(self):
        if not self.file_st.get():
            messagebox.showwarning("提示", "请先选择搜索词报告文件！")
            return

        # 更新阈值
        global ACOS_HIGH, ACOS_LOW, SPEND_THRESHOLD
        try:
            ACOS_HIGH = float(self.v_acos_high.get()) / 100
            ACOS_LOW = float(self.v_acos_low.get()) / 100
            SPEND_THRESHOLD = float(self.v_spend.get())
        except ValueError:
            messagebox.showerror("错误", "阈值请填写数字！")
            return

        self.btn.config(state='disabled')
        self.progress.start(10)
        threading.Thread(target=self._analyze, daemon=True).start()

    def _analyze(self):
        try:
            self._log("正在加载搜索词报告...")
            df_st = load_file(self.file_st.get())

            df_prod = None
            if self.file_prod.get():
                self._log("正在加载推广商品报告...")
                df_prod = load_file(self.file_prod.get())

            # 输出路径 = 搜索词报告同目录
            out_dir = os.path.dirname(self.file_st.get())
            timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            out_path = os.path.join(out_dir, f'广告诊断报告_{timestamp}.xlsx')

            result = build_report(df_st, df_prod, out_path,
                                  progress_callback=self._log)

            self.root.after(0, lambda: self._done(result, out_path))

        except Exception as e:
            self.root.after(0, lambda: self._error(str(e)))

    def _log(self, msg):
        self.root.after(0, lambda: self.status.config(text=msg))

    def _done(self, r, path):
        self.progress.stop()
        self.btn.config(state='normal')
        self.status.config(text="✅ 诊断完成！", fg='#006100')

        self.result_text.config(state='normal')
        self.result_text.delete('1.0', 'end')
        self.result_text.insert('end',
            f"分析完成！共 {r['total']} 个搜索词\n"
            f"─────────────────────────────\n"
            f"总花费: ${r['spend']:,.2f}  |  总销售: ${r['sales']:,.2f}  |  整体ACoS: {r['acos']:.1%}\n"
            f"🔴 烧钱词: {r['burn']}个  |  🟠 高ACoS: {r['high_acos']}个\n"
            f"🟢 优质词: {r['good']}个  |  🟡 潜力词: {r['potential']}个\n"
            f"💰 预估可节省: ${r['waste']:,.2f}\n"
            f"─────────────────────────────\n"
            f"报告已保存: {path}")
        self.result_text.config(state='disabled')

        if messagebox.askyesno("完成", f"诊断报告已生成！\n\n是否立即打开？"):
            os.startfile(path)

    def _error(self, msg):
        self.progress.stop()
        self.btn.config(state='normal')
        self.status.config(text="❌ 出错了", fg='CC0000')
        messagebox.showerror("分析出错", f"错误信息:\n{msg}\n\n请检查文件格式是否正确。")

    def run(self):
        self.root.mainloop()


if __name__ == '__main__':
    # 如果命令行传了参数，走无GUI模式
    if len(sys.argv) > 1:
        f_st = sys.argv[1]
        f_prod = sys.argv[2] if len(sys.argv) > 2 else None
        out = sys.argv[3] if len(sys.argv) > 3 else f_st.replace('.xlsx', '_诊断报告.xlsx').replace('.csv', '_诊断报告.xlsx')

        print(f"加载: {f_st}")
        df_st = load_file(f_st)
        df_prod = load_file(f_prod) if f_prod else None
        result = build_report(df_st, df_prod, out, progress_callback=print)
        print(f"\n完成！报告: {out}")
        print(f"烧钱词: {result['burn']}  优质词: {result['good']}  可节省: ${result['waste']:,.2f}")
    else:
        if HAS_TK:
            App().run()
        else:
            print("用法: python 亚马逊广告诊断工具.py <搜索词报告.xlsx> [推广商品报告.xlsx] [输出路径.xlsx]")
